"""Document ingestion service for various file formats"""
from pathlib import Path
import json
from typing import Optional
import logging

try:
    import pdfplumber
except Exception:  # pragma: no cover - optional dependency
    pdfplumber = None

try:
    from docx import Document as DocxDocument
except Exception:  # pragma: no cover - optional dependency
    DocxDocument = None

try:
    import openpyxl
except Exception:  # pragma: no cover - optional dependency
    openpyxl = None

try:
    import yaml
except Exception:  # pragma: no cover - optional dependency
    yaml = None

logger = logging.getLogger(__name__)

class DocumentIngestionService:
    """Handle document ingestion from various formats"""
    
    @staticmethod
    def ingest_pdf(file_path: str) -> str:
        """Extract text from PDF using pdfplumber"""
        try:
            if pdfplumber is None:
                raise RuntimeError("PDF extraction is unavailable because pdfplumber is not installed. Run: pip install pdfplumber")
            text = ""
            with pdfplumber.open(file_path) as pdf:
                for page_num, page in enumerate(pdf.pages):
                    extracted = page.extract_text()
                    if extracted:
                        text += f"\n[Page {page_num + 1}]\n{extracted}"
            if not text.strip():
                raise RuntimeError("No extractable text found in PDF. It may be scanned/image-only and needs OCR before ingestion.")
            logger.info(f"Successfully ingested PDF: {file_path}, length: {len(text)}")
            return text
        except Exception as e:
            logger.error(f"Error ingesting PDF {file_path}: {str(e)}")
            raise
    
    @staticmethod
    def ingest_docx(file_path: str) -> str:
        """Extract text from Word document"""
        try:
            if DocxDocument is None:
                raise RuntimeError("python-docx is not installed. Install requirements.txt to ingest DOCX files.")
            doc = DocxDocument(file_path)
            text = "\n".join([para.text for para in doc.paragraphs if para.text.strip()])
            logger.info(f"Successfully ingested DOCX: {file_path}, length: {len(text)}")
            return text
        except Exception as e:
            logger.error(f"Error ingesting DOCX {file_path}: {str(e)}")
            raise
    
    @staticmethod
    def ingest_txt(file_path: str) -> str:
        """Read plain text file"""
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                text = f.read()
            logger.info(f"Successfully ingested TXT: {file_path}, length: {len(text)}")
            return text
        except Exception as e:
            logger.error(f"Error ingesting TXT {file_path}: {str(e)}")
            raise
    
    @staticmethod
    def ingest_markdown(file_path: str) -> str:
        """Read markdown file"""
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                text = f.read()
            logger.info(f"Successfully ingested MD: {file_path}, length: {len(text)}")
            return text
        except Exception as e:
            logger.error(f"Error ingesting MD {file_path}: {str(e)}")
            raise
    
    @staticmethod
    def ingest_xlsx(file_path: str) -> str:
        """Extract text from Excel file"""
        try:
            if openpyxl is None:
                raise RuntimeError("openpyxl is not installed. Install requirements.txt to ingest Excel files.")
            text = ""
            workbook = openpyxl.load_workbook(file_path)
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text += f"\n[Sheet: {sheet_name}]\n"
                for row in sheet.iter_rows(values_only=True):
                    row_text = " | ".join([str(cell) if cell else "" for cell in row])
                    if row_text.strip():
                        text += row_text + "\n"
            logger.info(f"Successfully ingested XLSX: {file_path}, length: {len(text)}")
            return text
        except Exception as e:
            logger.error(f"Error ingesting XLSX {file_path}: {str(e)}")
            raise
    
    @staticmethod
    def ingest_json(file_path: str) -> str:
        """Extract text from JSON file"""
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            text = json.dumps(data, indent=2, ensure_ascii=False)
            logger.info(f"Successfully ingested JSON: {file_path}, length: {len(text)}")
            return text
        except Exception as e:
            logger.error(f"Error ingesting JSON {file_path}: {str(e)}")
            raise
    
    @staticmethod
    def ingest_yaml(file_path: str) -> str:
        """Extract text from YAML file"""
        try:
            if yaml is None:
                raise RuntimeError("pyyaml is not installed. Install requirements.txt to ingest YAML files.")
            with open(file_path, 'r', encoding='utf-8') as f:
                data = yaml.safe_load(f)
            text = yaml.dump(data, default_flow_style=False, allow_unicode=True)
            logger.info(f"Successfully ingested YAML: {file_path}, length: {len(text)}")
            return text
        except Exception as e:
            logger.error(f"Error ingesting YAML {file_path}: {str(e)}")
            raise
    
    @staticmethod
    def ingest(file_path: str) -> str:
        """Auto-detect format and ingest document"""
        ext = Path(file_path).suffix.lower()
        
        handlers = {
            '.pdf': DocumentIngestionService.ingest_pdf,
            '.docx': DocumentIngestionService.ingest_docx,
            '.doc': DocumentIngestionService.ingest_docx,
            '.txt': DocumentIngestionService.ingest_txt,
            '.md': DocumentIngestionService.ingest_markdown,
            '.xlsx': DocumentIngestionService.ingest_xlsx,
            '.xls': DocumentIngestionService.ingest_xlsx,
            '.json': DocumentIngestionService.ingest_json,
            '.yaml': DocumentIngestionService.ingest_yaml,
            '.yml': DocumentIngestionService.ingest_yaml,
        }
        
        if ext not in handlers:
            raise ValueError(f"Unsupported file format: {ext}")
        
        return handlers[ext](file_path)


DocumentService = DocumentIngestionService
